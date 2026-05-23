import io
import json

from flask import send_file
from flask_login import current_user

from app.extensions import db
from app.models import AuditLog


def log_activity(action, entity_type, entity_id=None, details=None):
    username = current_user.username if current_user.is_authenticated else 'system'
    log = AuditLog(
        user_id=current_user.id if current_user.is_authenticated else None,
        username=username,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=json.dumps(details or {}, ensure_ascii=False),
    )
    db.session.add(log)
    db.session.commit()


def export_workbook(headers, rows, filename):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename

    header_fill = PatternFill(start_color='4e73df', end_color='4e73df',
                              fill_type='solid')
    header_font = Font(color='ffffff', bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    ws.column_dimensions['A'].width = 8
    for col in ws.columns:
        if col[0].column_letter != 'A':
            ws.column_dimensions[col[0].column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=f'{filename}.xlsx', as_attachment=True)
